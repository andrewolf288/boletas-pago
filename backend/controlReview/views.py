from django.shortcuts import render
from django.contrib.auth.models import User, Group
from rest_framework import generics, permissions, viewsets, status
from rest_framework.decorators import action
from .models import *
from .serializers import *
from openpyxl import load_workbook
from rest_framework.response import Response
from django_user_agents.utils import get_user_agent
# from django.utils import timezone
from django.core.mail import EmailMessage
from django.core.mail import get_connection
from .utils import concatenar_mes_ano, convert_decimal_to_hours_minutes, excel_serial_to_date, send_email
import pandas as pd
from django.db import transaction
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils import timezone
import time
import random

import os

DOMAIN = 'http://localhost:5173'
EMAIL_TRANSMITTER = 'sistemas@emaransac.com'
class RemunerationViewSet(viewsets.ModelViewSet):
    queryset = Remuneration.objects.all()
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return RemunerationDetailSerializer
        return RemunerationSerializer

    def list(self, request):
        serializer = self.get_serializer(self.queryset, many=True)
        return Response(serializer.data)

    def retrieve(self, request, pk=None):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)
    
    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = request.data
        file = request.FILES.get('file')

        # si no se proporciono un archivo importado
        if not file:
            return Response({"detail": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            df = pd.read_excel(file, dtype=str)
            # primero debemos guardar la informacion de remuneración
            remuneration = Remuneration(
                year=data.get('year'),
                month=data.get('month'),
                duration=data.get('duration'),
                remunerationDateStart=data.get('remunerationDateStart'),
                remunerationDateEnd=data.get('remunerationDateEnd'),
                note=data.get('note')
            )
            remuneration.save()
            delay = 0

            for index, row in df.iterrows():
                # extraemos el valor de la primera columna (dcoument)
                extracted_data = row.iloc[0]

                # Verificar si el campo está vacío
                if pd.isna(extracted_data):
                    transaction.set_rollback(True)
                    return Response({"detail": f'Se encontró valor vacio en la primera columna de la fila {index+1}'}, status=status.HTTP_400_BAD_REQUEST)

                # Convertir a cadena y quitar espacios en blanco
                extracted_data = str(extracted_data).strip()
                # dato trabajador
                worker = Worker.objects.get(document=extracted_data)
                # informacion planilla
                voucher_data = {
                    'fechaInicioVacaciones': excel_serial_to_date(row.iloc[1]) if not pd.isna(row.iloc[1]) else '',
                    'fechaFinVacaciones': excel_serial_to_date(row.iloc[2]) if not pd.isna(row.iloc[2]) else '',
                    'diasLaborados': row.iloc[3] if not pd.isna(row.iloc[3]) else 0,
                    'diasNoLaborados': row.iloc[4] if not pd.isna(row.iloc[4]) else 0,
                    'horasLaboradas': convert_decimal_to_hours_minutes(row.iloc[5]) if not pd.isna(row.iloc[5]) else '',
                    'horasExtraSimples': convert_decimal_to_hours_minutes(row.iloc[6]) if not pd.isna(row.iloc[6]) else '',
                    'horasExtrasDobles': convert_decimal_to_hours_minutes(row.iloc[7]) if not pd.isna(row.iloc[7]) else '',
                    'bonificacionNocturna': row.iloc[8] if not pd.isna(row.iloc[8]) else 0,
                    'diasLicenciaGoceHaber': row.iloc[9] if not pd.isna(row.iloc[9]) else 0,
                    'diasFalta': row.iloc[10] if not pd.isna(row.iloc[10]) else 0,
                    'diasVacaciones': row.iloc[11] if not pd.isna(row.iloc[11]) else 0,
                    'diasDescansoMedico': row.iloc[12] if not pd.isna(row.iloc[12]) else 0,
                    'haberBasico': row.iloc[13] if not pd.isna(row.iloc[13]) else 0,
                    'asignacionFamiliar': row.iloc[14] if not pd.isna(row.iloc[14]) else 0,
                    'licenciaGoceHaber': row.iloc[15] if not pd.isna(row.iloc[15]) else 0,
                    'incapacidadEnfermedad': row.iloc[16] if not pd.isna(row.iloc[16]) else 0,
                    'hrsExtSimples25': row.iloc[17] if not pd.isna(row.iloc[17]) else 0,
                    'hrsExtSimples35': row.iloc[18] if not pd.isna(row.iloc[18]) else 0,
                    'diaDelTrabajador': row.iloc[19] if not pd.isna(row.iloc[19]) else 0,
                    'comisionDestajo': row.iloc[20] if not pd.isna(row.iloc[20]) else 0,
                    'gratificacion': row.iloc[21] if not pd.isna(row.iloc[21]) else 0,
                    'bonoExtraordinario': row.iloc[22] if not pd.isna(row.iloc[22]) else 0,
                    'cts': row.iloc[23] if not pd.isna(row.iloc[23]) else 0,
                    'vacaciones': row.iloc[24] if not pd.isna(row.iloc[24]) else 0,
                    'utilidad': row.iloc[25] if not pd.isna(row.iloc[25]) else 0,
                    'canastaVale': row.iloc[26] if not pd.isna(row.iloc[26]) else 0,
                    'premio': row.iloc[27] if not pd.isna(row.iloc[27]) else 0,
                    'snp': row.iloc[28] if not pd.isna(row.iloc[28]) else 0,
                    'afpFondo': row.iloc[29] if not pd.isna(row.iloc[29]) else 0,
                    'afpSeguro': row.iloc[30] if not pd.isna(row.iloc[30]) else 0,
                    'afpComision': row.iloc[31] if not pd.isna(row.iloc[31]) else 0,
                    'rtaStaCategoria': row.iloc[32] if not pd.isna(row.iloc[32]) else 0,
                    'adelantos': row.iloc[33] if not pd.isna(row.iloc[33]) else 0,
                    'esSaludVida': row.iloc[34] if not pd.isna(row.iloc[34]) else 0,
                    'tardanzaPermisosDescuentos': row.iloc[35] if not pd.isna(row.iloc[35]) else 0,
                    'inasistencia': row.iloc[36] if not pd.isna(row.iloc[36]) else 0,
                    'pagoGratificaciones': row.iloc[37] if not pd.isna(row.iloc[37]) else 0,
                    'pagoCTS': row.iloc[38] if not pd.isna(row.iloc[38]) else 0,
                    'pagoVacacionesBeneficios': row.iloc[39] if not pd.isna(row.iloc[39]) else 0,
                    'otrosDescuentos': row.iloc[40] if not pd.isna(row.iloc[40]) else 0,
                    'pagoUtilidad': row.iloc[41] if not pd.isna(row.iloc[41]) else 0,
                    'entCanastaVale': row.iloc[42] if not pd.isna(row.iloc[42]) else 0,
                    'esEssalud': row.iloc[43] if not pd.isna(row.iloc[43]) else 0,
                    'senati': row.iloc[44] if not pd.isna(row.iloc[44]) else 0,
                    'sctrSaludPension': row.iloc[45] if not pd.isna(row.iloc[45]) else 0,
                    'totalRemuneraciones': row.iloc[46] if not pd.isna(row.iloc[46]) else 0,
                    'totalDescuentos': row.iloc[47] if not pd.isna(row.iloc[47]) else 0,
                    'totalAportaciones': row.iloc[48] if not pd.isna(row.iloc[48]) else 0,
                    'netoPagar': row.iloc[49] if not pd.isna(row.iloc[49]) else 0,
                }
                voucher = Voucher(worker=worker, remuneration=remuneration, **voucher_data)
                voucher.save()

                # obtenemos el token del voucher
                voucher_token = voucher.token

                # parametros de email
                email_subject = f'BOLETA DE PAGO {str(remuneration.month).zfill(2)}-{str(remuneration.year)}'
                email_body = render_to_string('email_template.html', {
                    'date_voucher': concatenar_mes_ano(int(remuneration.month), remuneration.year),
                    'link_voucher': f'{DOMAIN}/voucherController/view?token={str(voucher_token)}',
                    'worker_fullname': f'{worker.user.first_name} {worker.user.last_name}',
                })
                email_plain_text = strip_tags(email_body)
                
                # enviamos email cada 5 a 10 segundos
                time.sleep(random.randint(5, 10))
                
                email_result = send_email(email_subject, email_plain_text, email_body, EMAIL_TRANSMITTER, worker.user.email)
                if email_result == 'Success':
                    voucher.sentEmail = True
                    voucher.sentEmailDate = timezone.now()
                else:
                    voucher.sentEmail = False
                    voucher.errorSend = email_result
                voucher.save()

                # send_email_task.apply_async(
                #     (voucher.id, email_subject, email_plain_text, email_body, EMAIL_TRANSMITTER, worker.user.email),
                #     countdown=delay
                # )
                # delay += 10
            
            return Response({"detail": "Información procesada y guardada correctamente"}, status=status.HTTP_201_CREATED)

        except Exception as e:
            transaction.set_rollback(True)
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class VoucherViewSet(viewsets.ViewSet):
    # permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def getVoucherByToken(self, request):
        token = request.data.get('token', None)
        document = request.data.get('document', None)

        if token is not None and document is not None:
            try:
                record = Voucher.objects.get(token=token)
                if record.worker.document == document:
                    if record.reviewed == True:
                        return Response({'detail': 'Recurso no disponible', 'data': record.reviewDate}, status=status.HTTP_410_GONE)
                    else:
                        serializer = VoucherSerializer(record)
                        return Response(serializer.data)
                else:
                    return Response({'detail': 'Las credenciales no corresponden'}, status=status.HTTP_401_UNAUTHORIZED)

            except Voucher.DoesNotExist:
                return Response({'detail': 'Record not found.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'detail': 'Token parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def verifyVoucherByToken(self, request):
        downloadPDF = request.data.get('downloadPDF', None)
        datetimeStartSession = request.data.get('datetimeStartSession', None)
        datetimeEndSession = request.data.get('datetimeEndSession', None)
        idVoucher = request.data.get('idVoucher', None)
        latitude = request.data.get('latitude', None)
        longitude = request.data.get('longitude', None)

        if idVoucher is not None:
            try:
                record = Voucher.objects.get(id=idVoucher)
                # Obtener información del dispositivo
                user_agent = get_user_agent(request)
                deviceType = 'PC' if user_agent.is_pc else 'Mobile' if user_agent.is_mobile else 'Tablet' if user_agent.is_tablet else 'Bot' if user_agent.is_bot else 'Unknown'

                # Obtener la dirección IP del cliente
                x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                if x_forwarded_for:
                    ip = x_forwarded_for.split(',')[0]
                else:
                    ip = request.META.get('REMOTE_ADDR')
                
                # Crear una nueva instancia de PaymentReceiptVerification
                verification = PaymentReceiptVerification(
                    voucher=record,
                    latitude=latitude,
                    longitude=longitude,
                    deviceType=deviceType,  # Puedes añadir lógica para obtener esto si es necesario
                    datetimeStartSession=datetimeStartSession,
                    datetimeEndSession=datetimeEndSession,
                    downloadPDF=downloadPDF,
                    ipAddress=ip,
                )
                verification.save()

                # Actualizar la instancia de Voucher
                record.reviewed = True
                print(timezone.now())
                record.reviewDate = timezone.now()
                record.save()

                return Response({'detail': 'Verification completed successfully.'}, status=status.HTTP_200_OK)

            except Voucher.DoesNotExist:
                return Response({'detail': 'Record not found.'}, status=status.HTTP_404_NOT_FOUND)   
        else:
            return Response({'detail': 'Parameters incorrects'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def forwardEmailVoucher(self, request):
        idVoucher = request.data.get('idVoucher', None)
        # if not request.user.is_authenticated:
        #     return Response({'detail': 'Authentication credentials were not provided.'}, status=status.HTTP_401_UNAUTHORIZED)

        if idVoucher is not None:
            try:
                voucher = Voucher.objects.get(id=idVoucher)
                voucher_token = voucher.token
                remuneration_id = voucher.remuneration.id
                remuneration = Remuneration.objects.get(id=remuneration_id)
                worker_id = voucher.worker.id
                worker = Worker.objects.get(id=worker_id)

                # parametros de email
                email_subject = f'BOLETA DE PAGO {str(remuneration.month).zfill(2)}-{str(remuneration.year)}'
                email_body = render_to_string('email_template.html', {
                    'date_voucher': concatenar_mes_ano(int(remuneration.month), remuneration.year),
                    'link_voucher': f'http://localhost:5173/voucherController/view?token={str(voucher_token)}'
                })
                email_plain_text = strip_tags(email_body)
                email_result = send_email(email_subject, email_plain_text, email_body, EMAIL_TRANSMITTER, worker.user.email)
                
                if email_result == 'Success':
                    voucher.sentEmail = True
                    voucher.sentEmailDate = timezone.now()
                    voucher.save()
                    return Response({'detail': 'Correo electrónico enviado.'}, status=status.HTTP_200_OK)
                else:
                    voucher.sentEmail = False
                    voucher.errorSend = email_result
                    voucher.save()
                    return Response({'detail': email_result}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
            except Voucher.DoesNotExist:
                return Response({'detail': 'Registro no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'detail': 'Parametros incorrectos'}, status=status.HTTP_400_BAD_REQUEST)

# Funcion de importación de trabajadores
def cargar_datos_desde_excel(request):
    grupo_trabajador, created = Group.objects.get_or_create(name='Trabajador')

    ruta_archivo = os.path.join(os.path.dirname(__file__), 'data', 'data.xlsx')
    workbook = load_workbook(filename=ruta_archivo)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        first_name = row[0]
        last_name = row[1]
        email = row[2]
        username = row[3]
        password = row[4]

        user = User.objects.create_user(
            username = username,
            password = password,
            first_name = first_name,
            last_name = last_name,
            email = email,
        )

        user.groups.add(grupo_trabajador)

        typeDocument = row[6]
        type_document = TypeDocument.objects.get(pk=typeDocument)
        document = row[7]
        workRegime = row[8]
        typeWorker = row[9]
        type_worker = TypeWorker.objects.get(pk=typeWorker)
        sede = row[16]
        situation = row[18]

        worker = Worker.objects.create(
            user=user,
            typeDocument=type_document,
            typeWorker=type_worker,
            document=document,
            workRegime=workRegime,
            sede=sede,
            situation=situation
        )